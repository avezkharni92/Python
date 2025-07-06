import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# File path
FILE_NAME = "study_tracker.xlsx"

# Check if file exists; if not, create with headers
if not os.path.isfile(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.title = "Study Log"
    ws.append(["Date", "Time Spent", "Topic", "Streak"])
    wb.save(FILE_NAME)
    print("Created new Excel tracker.")

# Load workbook
wb = load_workbook(FILE_NAME)
ws = wb.active

# Ask user for input
date = input("Enter Date (DD/MM/YYYY) [default: today]: ").strip()
if not date:
    date = datetime.now().strftime("%d/%m/%Y")

time_spent = input("Enter Time Spent (e.g., 0h:25m:12s): ").strip()
topic = input("Enter Topic Studied: ").strip()
streak = input("Enter Streak Info (e.g., Day 5): ").strip()

# Append new row
ws.append([date, time_spent, topic, streak])
wb.save(FILE_NAME)

print("✅ Study session saved successfully to Excel!")
# Close workbook
wb.close()